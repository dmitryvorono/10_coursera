import requests
from lxml import etree
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook
import argparse
import random


def fetch_content(url):
    request = requests.get(url)
    if request.status_code == requests.codes.ok:
        return request.content


def get_courses_list(coursera_xml):
    coursera_tree_xml = etree.XML(coursera_xml)
    return [e.text for e in coursera_tree_xml.iter('*') if e.text.startswith('http')]


def get_html_text(html_tree, tag, attrs):
    target_element = html_tree.find(tag, attrs)
    if target_element is not None:
        return target_element.text


def get_html_count_elements(html_tree, tag, attrs):
    target_elements = html_tree.find_all(tag, attrs)
    return len(target_elements)


def parse_course(course_html):
    course_html_tree = BeautifulSoup(course_html, 'html.parser')
    return {'Name': get_html_text(course_html_tree, 'h1', {'class': 'title display-3-text'}),
            'Language': get_html_text(course_html_tree, 'div', {'class': 'rc-Language'}),
            'Start date': course_html_tree.find('div', {'class': 'startdate rc-StartDateString caption-text'}).find('span').text,
            'Period': get_html_count_elements(course_html_tree, 'div', {'class': 'week'}),
            'Rating': get_html_text(course_html_tree, 'div', {'class': 'ratings-text bt3-visible-xs'})
            }


def get_course_info(course_url):
    course_html = fetch_content(course_url)
    return(parse_course(course_html))


def xls_set_head(worksheet, head_names):
    for row in worksheet.iter_rows(min_row=1, max_col=len(head_names) + 1, max_row=1):
        for name, cell in zip(head_names, row):
            cell.value = name


def format_worksheet(worksheet, title, head_names):
    worksheet.title = title
    xls_set_head(worksheet, head_names)


def output_courses_info_to_xlsx(courses_info):
    wb = Workbook()
    worksheet = wb.active
    course_keys = courses_info[0].keys()
    format_worksheet(worksheet, 'Coursera', course_keys)
    for row, course_info in zip(worksheet.iter_rows(min_row=2, max_col=len(course_keys)+1, max_row=len(courses_info)+1), courses_info):
        for name, cell in zip(course_keys, row):
            cell.value = course_info[name]
    return wb


def get_command_line_arguments():
    default_count_courses = 20
    default_output_filename = 'coursera_courses.xlsx'
    parser = argparse.ArgumentParser(description='export Cousera courses to excel')
    parser.add_argument('-c', '--count', help='count courses to export in excel', type=int, default=default_count_courses)
    parser.add_argument('-o', '--output', help='output filename', type=str, default=default_output_filename)
    return parser.parse_args()


if __name__ == '__main__':
    args = get_command_line_arguments()
    coursera_xml_feed = 'https://www.coursera.org/sitemap~www~courses.xml'
    coursera_xml_content = fetch_content(coursera_xml_feed)
    random_courses_list = random.sample(get_courses_list(coursera_xml_content), args.count)
    courses_info = []
    for course_url in random_courses_list:
        print('fetch: {0}'.format(course_url))
        courses_info.append(get_course_info(course_url))
        time.sleep(1)
    excel_workbook = output_courses_info_to_xlsx(courses_info)
    excel_workbook.save(args.output)
